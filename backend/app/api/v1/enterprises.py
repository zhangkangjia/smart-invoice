"""企业管理路由。"""

import uuid
from datetime import datetime, timezone
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.deps import get_current_active_user
from app.db.session import get_db
from app.models.enterprise import Enterprise, EnterpriseConfig, ServiceAssignment
from app.models.user import User
from app.schemas.common import PaginatedResponse
from app.schemas.enterprise import (
    EnterpriseBrief,
    EnterpriseConfigResponse,
    EnterpriseConfigUpdate,
    EnterpriseCreate,
    EnterpriseResponse,
    EnterpriseUpdate,
    ServiceAssignmentCreate,
    ServiceAssignmentResponse,
)
from app.services import enterprise_service
from app.services.audit_service import log_action

router = APIRouter(prefix="/enterprises", tags=["企业管理"])


async def _get_enterprise_or_404(
    db: AsyncSession,
    tenant_id: str,
    enterprise_id: str,
) -> Enterprise:
    result = await db.execute(
        select(Enterprise).where(
            Enterprise.id == enterprise_id,
            Enterprise.tenant_id == tenant_id,
        )
    )
    enterprise = result.scalar_one_or_none()
    if enterprise is None:
        raise HTTPException(status_code=404, detail="企业不存在")
    return enterprise


@router.get("", response_model=PaginatedResponse[EnterpriseResponse], summary="企业列表")
async def list_enterprises(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    status_filter: str | None = Query(None, alias="status"),
    keyword: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """分页查询企业列表。"""
    enterprises, total = await enterprise_service.get_enterprises(
        db,
        tenant_id=current_user.tenant_id,
        page=page,
        page_size=page_size,
        status_filter=status_filter,
        keyword=keyword,
    )
    items = [EnterpriseResponse.model_validate(e) for e in enterprises]
    return PaginatedResponse.create(items, total, page, page_size)


@router.get("/lookup/by-tax-no", summary="通过税号查询企业信息")
async def lookup_by_tax_no(
    tax_no: str = Query(..., min_length=6, max_length=20, description="纳税人识别号"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """通过税号查询企业信息。

    1. 先查本地数据库是否已存在该企业
    2. 若不存在，尝试通过百望云通道查询企业基本信息
    3. 百望云未配置时返回税号格式校验结果（可手动补全）
    """
    import re
    import logging
    logger = logging.getLogger(__name__)

    # 税号格式校验
    tax_no_clean = tax_no.strip().upper()
    if not re.match(r'^[0-9A-Z]{15,20}$', tax_no_clean):
        raise HTTPException(status_code=400, detail="税号格式不正确（应为15-20位字母数字）")

    result_data: dict = {
        "tax_no": tax_no_clean,
        "found_locally": False,
        "found_remote": False,
        "enterprise": None,
        "suggestion": None,
        "message": "",
    }

    # 1. 查本地
    local_result = await db.execute(
        select(Enterprise).where(
            Enterprise.tenant_id == current_user.tenant_id,
            Enterprise.tax_no == tax_no_clean,
        )
    )
    local_enterprise = local_result.scalar_one_or_none()
    if local_enterprise:
        result_data["found_locally"] = True
        result_data["enterprise"] = EnterpriseResponse.model_validate(local_enterprise).model_dump()
        result_data["message"] = "该企业已在系统中存在"
        return result_data

    # 2. 尝试百望云查询
    try:
        from app.services.channels.registry import ChannelRegistry
        bw_channel = ChannelRegistry.get("baiwang")
        if bw_channel and hasattr(bw_channel, 'config') and bw_channel.config.app_key:
            remote_info = await bw_channel.query_enterprise_info(tax_no_clean)
            if remote_info:
                result_data["found_remote"] = True
                result_data["suggestion"] = {
                    "name": remote_info.get("nsrmc", ""),
                    "tax_no": tax_no_clean,
                    "address": remote_info.get("scjydz", ""),
                    "phone": remote_info.get("lxdh", ""),
                    "bank_name": remote_info.get("khyh", ""),
                    "bank_account": remote_info.get("yhzh", ""),
                }
                result_data["message"] = "已从百望云获取企业信息，请确认后保存"
                return result_data
    except Exception as e:
        logger.warning("百望云企业信息查询失败: %s", e)

    # 3. 百望云未配置或查询失败
    result_data["message"] = "未配置百望云通道或查询失败，请手动填写企业信息"
    return result_data


@router.get("/template/download", summary="下载企业导入模板")
async def download_template(
    current_user: User = Depends(get_current_active_user),
):
    """下载企业批量导入 Excel 模板。"""
    import openpyxl
    from urllib.parse import quote

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "企业导入模板"

    # 表头
    headers = [
        "企业名称(必填)", "税号(必填)", "地址", "电话",
        "开户银行", "银行账号", "状态(填active或留空)",
    ]
    ws.append(headers)

    # 示例行
    ws.append([
        "杭州两心同网络科技有限公司",
        "91330106MA28T1234X",
        "浙江省杭州市西湖区文三路478号",
        "13800138000",
        "中国工商银行杭州西湖支行",
        "1202020809000123456",
        "active",
    ])

    # 列宽
    widths = [32, 22, 40, 16, 28, 24, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 说明行
    ws.append([])
    ws.append(["说明:"])
    ws.append(["1. 企业名称和税号为必填，其余可选"])
    ws.append(["2. 状态留空默认为 active（启用），可填 active/pending/suspended"])
    ws.append(["3. 税号重复的行将自动跳过（按已有税号更新）"])
    ws.append(["4. 税号格式：15-20位字母数字"])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = "企业导入模板.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"
        },
    )


@router.post("/import", summary="批量导入企业")
async def import_enterprises(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """通过 Excel 批量导入企业。

    模板列: 企业名称 | 税号 | 地址 | 电话 | 开户银行 | 银行账号 | 状态
    - 税号已存在则更新，不存在则新增
    - 返回成功/失败/跳过数量及错误明细
    """
    import openpyxl

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传 Excel 文件(.xlsx)")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel 文件解析失败: {e}")

    ws = wb.active

    # 读取所有行（跳过表头）
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    success_count = 0
    update_count = 0
    skip_count = 0
    errors: list[dict] = []

    for idx, row in enumerate(rows, start=2):
        # 跳过空行
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            skip_count += 1
            continue

        name = str(row[0]).strip() if row[0] else ""
        tax_no = str(row[1]).strip().upper() if row[1] else ""

        if not name:
            errors.append({"row": idx, "error": "企业名称为空"})
            continue
        if not tax_no:
            errors.append({"row": idx, "error": "税号为空"})
            continue

        address = str(row[2]).strip() if len(row) > 2 and row[2] else None
        phone = str(row[3]).strip() if len(row) > 3 and row[3] else None
        bank_name = str(row[4]).strip() if len(row) > 4 and row[4] else None
        bank_account = str(row[5]).strip() if len(row) > 5 and row[5] else None
        ent_status = str(row[6]).strip().lower() if len(row) > 6 and row[6] else "active"

        if ent_status not in ("active", "pending", "suspended", "archived"):
            ent_status = "active"

        # 查是否已存在
        existing = await db.execute(
            select(Enterprise).where(
                Enterprise.tenant_id == current_user.tenant_id,
                Enterprise.tax_no == tax_no,
            )
        )
        existing_ent = existing.scalar_one_or_none()

        if existing_ent:
            # 更新
            existing_ent.name = name
            if address: existing_ent.address = address
            if phone: existing_ent.phone = phone
            if bank_name: existing_ent.bank_name = bank_name
            if bank_account: existing_ent.bank_account = bank_account
            existing_ent.status = ent_status
            update_count += 1
        else:
            # 新增
            enterprise = Enterprise(
                tenant_id=current_user.tenant_id,
                name=name,
                tax_no=tax_no,
                address=address,
                phone=phone,
                bank_name=bank_name,
                bank_account=bank_account,
                status=ent_status,
            )
            db.add(enterprise)
            success_count += 1

    try:
        await db.commit()
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"数据库保存失败: {e}")

    return {
        "total": len(rows),
        "created": success_count,
        "updated": update_count,
        "skipped": skip_count,
        "errors": errors,
    }


@router.post("", response_model=EnterpriseResponse, status_code=status.HTTP_201_CREATED, summary="创建企业")
async def create_enterprise(
    body: EnterpriseCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """创建企业。"""
    data = body.model_dump()
    enterprise = await enterprise_service.create_enterprise(
        db,
        tenant_id=current_user.tenant_id,
        data=data,
        created_by=current_user.id,
    )
    await log_action(
        db=db,
        tenant_id=current_user.tenant_id,
        user_id=current_user.id,
        action="create_enterprise",
        entity_type="enterprise",
        entity_id=enterprise.id,
        after={"name": enterprise.name, "tax_no": enterprise.tax_no},
    )
    await db.commit()
    return EnterpriseResponse.model_validate(enterprise)


@router.get("/{enterprise_id}", response_model=EnterpriseResponse, summary="企业详情")
async def get_enterprise(
    enterprise_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """获取企业详情。"""
    enterprise = await _get_enterprise_or_404(db, current_user.tenant_id, enterprise_id)
    return EnterpriseResponse.model_validate(enterprise)


@router.put("/{enterprise_id}", response_model=EnterpriseResponse, summary="更新企业")
async def update_enterprise(
    enterprise_id: str,
    body: EnterpriseUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """更新企业基本信息。"""
    enterprise = await _get_enterprise_or_404(db, current_user.tenant_id, enterprise_id)
    await enterprise_service.update_enterprise(db, enterprise, body.model_dump(exclude_unset=True))
    await log_action(
        db=db,
        tenant_id=current_user.tenant_id,
        user_id=current_user.id,
        action="update_enterprise",
        entity_type="enterprise",
        entity_id=enterprise.id,
    )
    await db.commit()
    return EnterpriseResponse.model_validate(enterprise)


@router.patch("/{enterprise_id}/status", response_model=EnterpriseResponse, summary="更新企业状态")
async def update_enterprise_status(
    enterprise_id: str,
    new_status: str = Query(..., alias="status"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """更新企业状态（带状态机校验）。"""
    enterprise = await _get_enterprise_or_404(db, current_user.tenant_id, enterprise_id)
    old_status = enterprise.status
    await enterprise_service.update_enterprise_status(db, enterprise, new_status)
    await log_action(
        db=db,
        tenant_id=current_user.tenant_id,
        user_id=current_user.id,
        action="update_enterprise_status",
        entity_type="enterprise",
        entity_id=enterprise.id,
        before={"status": old_status},
        after={"status": new_status},
    )
    await db.commit()
    return EnterpriseResponse.model_validate(enterprise)


@router.get("/{enterprise_id}/config", response_model=EnterpriseConfigResponse, summary="企业配置")
async def get_enterprise_config(
    enterprise_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """获取企业开票配置。"""
    await _get_enterprise_or_404(db, current_user.tenant_id, enterprise_id)
    config = await enterprise_service.get_or_create_config(
        db, current_user.tenant_id, enterprise_id
    )
    await db.commit()
    return EnterpriseConfigResponse.model_validate(config)


@router.put("/{enterprise_id}/config", response_model=EnterpriseConfigResponse, summary="更新配置")
async def update_enterprise_config(
    enterprise_id: str,
    body: EnterpriseConfigUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """更新企业开票配置。"""
    await _get_enterprise_or_404(db, current_user.tenant_id, enterprise_id)
    config = await enterprise_service.get_or_create_config(
        db, current_user.tenant_id, enterprise_id
    )
    await enterprise_service.update_enterprise_config(
        db, config, body.model_dump(exclude_unset=True)
    )
    await log_action(
        db=db,
        tenant_id=current_user.tenant_id,
        user_id=current_user.id,
        action="update_enterprise_config",
        entity_type="enterprise_config",
        entity_id=config.id,
    )
    await db.commit()
    return EnterpriseConfigResponse.model_validate(config)


@router.get(
    "/{enterprise_id}/assignments",
    response_model=list[ServiceAssignmentResponse],
    summary="服务人员列表",
)
async def get_assignments(
    enterprise_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """获取企业服务人员列表。"""
    await _get_enterprise_or_404(db, current_user.tenant_id, enterprise_id)
    assignments = await enterprise_service.get_service_assignments(
        db, current_user.tenant_id, enterprise_id
    )
    return [ServiceAssignmentResponse.model_validate(a) for a in assignments]


@router.post(
    "/{enterprise_id}/assignments",
    response_model=ServiceAssignmentResponse,
    status_code=status.HTTP_201_CREATED,
    summary="分配服务人员",
)
async def assign_service_person(
    enterprise_id: str,
    body: ServiceAssignmentCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """为企业分配服务人员。"""
    await _get_enterprise_or_404(db, current_user.tenant_id, enterprise_id)
    assignment = await enterprise_service.assign_service_person(
        db,
        tenant_id=current_user.tenant_id,
        enterprise_id=enterprise_id,
        user_id=body.user_id,
        role=body.role,
        assigned_by=current_user.id,
        start_at=body.start_at,
        end_at=body.end_at,
    )
    await log_action(
        db=db,
        tenant_id=current_user.tenant_id,
        user_id=current_user.id,
        action="assign_service_person",
        entity_type="service_assignment",
        entity_id=assignment.id,
        after={"user_id": body.user_id, "role": body.role},
    )
    await db.commit()
    return ServiceAssignmentResponse.model_validate(assignment)


@router.get("/{enterprise_id}/health", summary="企业健康度")
async def get_enterprise_health(
    enterprise_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """计算并返回企业健康度评分。"""
    await _get_enterprise_or_404(db, current_user.tenant_id, enterprise_id)
    health = await enterprise_service.calculate_enterprise_health(
        db, current_user.tenant_id, enterprise_id
    )
    return health
