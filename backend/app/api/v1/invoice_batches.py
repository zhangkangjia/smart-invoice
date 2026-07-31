"""开票批次路由。"""

from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.deps import get_current_active_user
from app.db.session import get_db
from app.models.enterprise import Enterprise
from app.models.invoice import InvoiceRequest
from app.models.invoice_task import InvoiceResult, InvoiceTask
from app.models.task import ImportBatch
from app.models.user import User
from app.schemas.common import PaginatedResponse
from app.schemas.task import ImportBatchResponse

router = APIRouter(prefix="/invoice-batches", tags=["开票批次"])


@router.get("/template", summary="下载Excel开票模板")
async def download_excel_template(
    current_user: User = Depends(get_current_active_user),
):
    """下载标准批量开票Excel模板。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "批量开票"
    headers = [
        "企业名称", "购方名称", "购方税号", "商品名称", "规格", "单位",
        "数量", "单价", "税率", "接收邮箱", "备注",
    ]
    worksheet.append(headers)
    worksheet.append([
        "示例销方企业", "示例购方企业", "", "技术服务费", "", "次",
        1, 10000, 6, "finance@example.com", "",
    ])
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
    for column, width in {
        "A": 24, "B": 28, "C": 22, "D": 22, "E": 14, "F": 10,
        "G": 10, "H": 14, "I": 10, "J": 28, "K": 24,
    }.items():
        worksheet.column_dimensions[column].width = width
    worksheet.freeze_panes = "A2"

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    # HTTP Header 必须使用 latin-1；使用 ASCII fallback 避免中文文件名导致响应500。
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=invoice-import-template.xlsx; filename*=UTF-8''%E6%89%B9%E9%87%8F%E5%BC%80%E7%A5%A8%E5%AF%BC%E5%85%A5%E6%A8%A1%E6%9D%BF.xlsx"
        },
    )


@router.get("", response_model=PaginatedResponse[ImportBatchResponse], summary="批次列表")
async def list_batches(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """分页查询导入批次列表。"""
    stmt = select(ImportBatch).where(ImportBatch.tenant_id == current_user.tenant_id)
    total = (await db.execute(select(func.count()).select_from(stmt.subquery()))).scalar_one()
    offset = (page - 1) * page_size
    result = await db.execute(
        stmt.order_by(ImportBatch.created_at.desc()).offset(offset).limit(page_size)
    )
    items = [ImportBatchResponse.model_validate(b) for b in result.scalars().all()]
    return PaginatedResponse.create(items, total, page, page_size)


@router.get("/{batch_id}", response_model=ImportBatchResponse, summary="批次详情")
async def get_batch(
    batch_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """获取批次详情。"""
    result = await db.execute(
        select(ImportBatch).where(
            ImportBatch.id == batch_id,
            ImportBatch.tenant_id == current_user.tenant_id,
        )
    )
    batch = result.scalar_one_or_none()
    if batch is None:
        raise HTTPException(status_code=404, detail="批次不存在")
    return ImportBatchResponse.model_validate(batch)


@router.get("/{batch_id}/tasks", summary="批次下任务列表")
async def list_batch_tasks(
    batch_id: str,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    status_filter: str | None = Query(None, alias="status"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """获取批次下的开票任务列表（含企业、购方、金额、发票号等关联数据）。"""
    stmt = select(InvoiceTask).where(
        InvoiceTask.tenant_id == current_user.tenant_id,
        InvoiceTask.import_batch_id == batch_id,
    )
    if status_filter:
        stmt = stmt.where(InvoiceTask.status == status_filter)

    total = (await db.execute(select(func.count()).select_from(stmt.subquery()))).scalar_one()
    offset = (page - 1) * page_size
    result = await db.execute(
        stmt.order_by(InvoiceTask.created_at.asc()).offset(offset).limit(page_size)
    )
    tasks = result.scalars().all()

    enterprise_ids = list({t.enterprise_id for t in tasks})
    request_ids = list({t.invoice_request_id for t in tasks})
    task_ids = [t.id for t in tasks]

    enterprises: dict = {}
    requests: dict = {}
    results: dict = {}

    if enterprise_ids:
        rows = await db.execute(select(Enterprise).where(Enterprise.id.in_(enterprise_ids)))
        enterprises = {r.id: r for r in rows.scalars().all()}
    if request_ids:
        rows = await db.execute(select(InvoiceRequest).where(InvoiceRequest.id.in_(request_ids)))
        requests = {r.id: r for r in rows.scalars().all()}
    if task_ids:
        rows = await db.execute(select(InvoiceResult).where(InvoiceResult.invoice_task_id.in_(task_ids)))
        results = {r.invoice_task_id: r for r in rows.scalars().all()}

    return {
        "items": [
            {
                "id": t.id,
                "status": t.status,
                "enterprise_id": t.enterprise_id,
                "enterprise_name": enterprises.get(t.enterprise_id).name if enterprises.get(t.enterprise_id) else None,
                "invoice_request_id": t.invoice_request_id,
                "buyer_name": requests.get(t.invoice_request_id).buyer_name if requests.get(t.invoice_request_id) else None,
                "buyer_tax_no": requests.get(t.invoice_request_id).buyer_tax_no if requests.get(t.invoice_request_id) else None,
                "total_with_tax": float(requests.get(t.invoice_request_id).total_with_tax or 0) if requests.get(t.invoice_request_id) else 0,
                "invoice_number": results.get(t.id).invoice_number if results.get(t.id) else None,
                "retry_count": t.retry_count,
                "last_error": t.last_error,
                "created_at": t.created_at.isoformat() if t.created_at else None,
                "completed_at": t.completed_at.isoformat() if t.completed_at else None,
            }
            for t in tasks
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.get("/{batch_id}/failed-rows/export", summary="导出失败行Excel")
async def export_failed_rows(
    batch_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """导出指定批次中失败或异常的开票任务明细，供用户修正后重新导入。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    stmt = select(InvoiceTask).where(
        InvoiceTask.tenant_id == current_user.tenant_id,
        InvoiceTask.import_batch_id == batch_id,
        InvoiceTask.status.in_(["failed", "unknown", "awaiting_manual", "awaiting_reconciliation"]),
    )
    tasks = (await db.execute(stmt)).scalars().all()

    request_ids = [t.invoice_request_id for t in tasks]
    requests_map: dict = {}
    if request_ids:
        rows = await db.execute(select(InvoiceRequest).where(InvoiceRequest.id.in_(request_ids)))
        requests_map = {r.id: r for r in rows.scalars().all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "失败任务"
    headers = ["购方名称", "购方税号", "商品名称", "数量", "单价", "税率", "价税合计", "失败原因", "任务状态"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="DC2626")

    for task in tasks:
        req = requests_map.get(task.invoice_request_id)
        ws.append([
            req.buyer_name if req else "",
            req.buyer_tax_no if req else "",
            "",  # 商品名称需查item，此处简化
            float(req.total_with_tax) if req else 0,
            "",
            "",
            float(req.total_with_tax) if req else 0,
            task.last_error or "",
            task.status,
        ])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=failed-rows.xlsx"},
    )
